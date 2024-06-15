import RPi.GPIO as GPIO
from piservo import Servo
import thingspeak
import time
import openpyxl
import re
import datetime
from picamera import PiCamera

import cv2
import imutils
import pytesseract

wb=openpyxl.load_workbook("/home/pi/Desktop/codes/slot_booking_details.xlsx")
sheet=wb.active

wb1=openpyxl.load_workbook("/home/pi/Desktop/codes/Vehicle_Number(Not_booked_slot).xlsx")
sheet1=wb1.active


time.sleep(3)
camera = PiCamera()

channel_id=1833772
write_key='X4H7ZRBOYPAHYP9M'

GPIO.setmode(GPIO.BCM)
GPIO.setwarnings(False)

# Allocation of GPIO PINS...
###################################
#1.For IR Sensor 1
ir1= 20
GPIO.setup(ir1,GPIO.IN)

#1.1 For IR Sensor 2
ir2=21
GPIO.setup(ir2,GPIO.IN)

###################################

#2. For Ultrasonic sensor-1
slot1_echo= 14
slot1_trig=15
GPIO.setup(slot1_echo,GPIO.IN)
GPIO.setup(slot1_trig,GPIO.OUT)

#2.1 For Ultrasonic sensor -2

slot2_echo= 23
slot2_trig=24
GPIO.setup(slot2_echo,GPIO.IN)
GPIO.setup(slot2_trig,GPIO.OUT)

######################################

#For Servo Motor(Gate)
myservo = Servo(19)

# GPIO.setup(19, GPIO.OUT)
# servo = GPIO.PWM(19, 50)
# servo.start(0)

#setting initial position
# Set the initial position to 0 degrees
# position = 0
# servo.ChangeDutyCycle(2 + (position / 18))
# 
# # Rotate the servo motor 90 degrees forward
# def servo_open():
#     position = 90
#     servo.ChangeDutyCycle(2 + (position / 18))
#     servo.stop()
# def servo_close():
# # Rotate the servo motor 90 degrees backward
#      position = 0
#      servo.ChangeDutyCycle(2 + (position / 18))




######################################

def car_detection(TRIGGER,ECHO):
    # set Trigger to HIGH
    GPIO.output(TRIGGER, True)
 
    # set Trigger after 0.01ms to LOW
    time.sleep(0.00001)
    GPIO.output(TRIGGER, False)
    StartTime=0
    StopTime=0
    # save StartTime
    while GPIO.input(ECHO) == 0:
        StartTime = time.time()
 
    # save time of arrival
    while GPIO.input(ECHO) == 1:
        StopTime = time.time()
 
    # time difference between start and arrival
    TimeElapsed = StopTime-StartTime
    # multiply with the sonic speed (34300 cm/s)
    # and divide by 2, because there and back
    distance = (TimeElapsed * 34300) / 2
 
    return distance

#############################################
def cloud(slot1,slot2,count):
    #response = channel.update('field1': slot1,'field2': slot2,'field3': count)
    #response = channel.update({"Field3": count})
    #response = channel.update({"Field2": slot2})
    #print("Slot-2 sent")
    #response = channel.update({"Field1": slot1})
    #print("Slot-1 sent")
    #response = channel.update({"Field3": count})
    #print("Count sent")
    client=thingspeak.Channel(channel_id, write_key)
    data={
        "field1":slot1,
        "field2":slot2,
        "field3":count,
        
        }
    response=client.update(data)
    print("Data sent to cloud...")
    
#############################################
    
def find_and_insert(plate):
    i=1
    flag=0
    a=sheet.cell(row=1,column=2)
    while(a.value != None):
        a=sheet.cell(row=i,column=2)
        if(a.value==plate):
            b=sheet.cell(row=i,column=3)
            if(b.value== datetime.date.today()):
                flag=1
                break
        i=i+1
    if(flag==1):
        print("This car has a pre-booking...")
    else:
        a=sheet1.cell(row=1,column=1)
        i=1
        while(a.value!=None):
            a=sheet1.cell(row=i,column=1)
            i=i+1
        a.value=plate
        a=sheet1.cell(row=i,column=2)
        a.value=datetime.datetime.now()
        wb1.save("/home/pi/Desktop/codes/Vehicle_Number(Not_booked_slot).xlsx")
        print("This car has no pre-booking....")
        
        
def numberplate_detect():
    
    camera.capture('number_plate.jpg')
    image = cv2.imread('/home/pi/Desktop/codes/cars.jpg')
    image = imutils.resize(image, width=300 )
    #cv2.imshow("original image", image)
    #cv2.waitKey(0)

    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    #cv2.imshow("greyed image", gray_image)
    #cv2.waitKey(0)

    gray_image = cv2.bilateralFilter(gray_image, 11, 17, 17) 
    #cv2.imshow("smoothened image", gray_image)
    #cv2.waitKey(0)

    edged = cv2.Canny(gray_image, 30, 200) 
    #cv2.imshow("edged image", edged)
    #cv2.waitKey(0)

    cnts,new = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    image1=image.copy()
    cv2.drawContours(image1,cnts,-1,(0,255,0),3)
    #cv2.imshow("contours",image1)
    #cv2.waitKey(0)

    cnts = sorted(cnts, key = cv2.contourArea, reverse = True) [:30]
    screenCnt = None
    image2 = image.copy()
    cv2.drawContours(image2,cnts,-1,(0,255,0),3)
    #cv2.imshow("Top 30 contours",image2)
    #cv2.waitKey(0)

    i=7
    for c in cnts:
        perimeter = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.018 * perimeter, True)
        if len(approx) == 4: 
                screenCnt = approx
                
                x,y,w,h = cv2.boundingRect(c) 
                new_img=image[y:y+h,x:x+w]
                cv2.imwrite('./'+str(i)+'.png',new_img)
                i+=1
                break
            
    cv2.drawContours(image, [screenCnt], -1, (0, 255, 0), 3)
    #cv2.imshow("image with detected license plate", image)
    cv2.waitKey(0)

    Cropped_loc = './7.png'
    #cv2.imshow("cropped", cv2.imread(Cropped_loc))
    plate = pytesseract.image_to_string(Cropped_loc, lang='eng')
    print("Number plate is:", plate)
    #print(type(plate))
    #print(len(plate))
    
    pattern = re.compile('[\W_]+')
    num_plate = pattern.sub('', plate)
    print(num_plate)
    #print(len(num_plate))

    
    find_and_insert(num_plate)
    #cv2.waitKey(0)
    cv2.destroyAllWindows()
    
    
############################################
    
    
if __name__ == "__main__":
    count=0
    #channel = thingspeak.Channel(id=channel_id, api_key=write_keyID)
    while True:
        
        if (GPIO.input(ir1)==0) and count<2:
            #servo_open()
            print("Detecting Number plate")
            numberplate_detect()
            myservo.write(90)
            print("Gate Opened")
            time.sleep(3)
            while GPIO.input(ir2) == 1:
                continue
            #servo_close()
            myservo.write(0)
            print("Gate Closed")
            time.sleep(3)
            count=count+1
            dist1=car_detection(slot1_trig,slot1_echo)
            dist2=car_detection(slot2_trig,slot2_echo)
            if (dist1<50):
                print("Car parked in slot-1")
                print(dist1)
                a=1
            else:
                print("No car parked in slot-1")
                print(dist1)
                a=0
            if (dist2<50):
                print("Car parked in slot-2")
                print(dist2)
                b=1
            else:
                print("No car parked in slot-2")
                print(dist2)
                b=0
            cloud(a,b,count)
            print("Number of cars parked =",count)
            time.sleep(15)
                
        if GPIO.input(ir2)==0 and count>0:
            #servo_open()
            myservo.write(90)
            print("Gate Opened")
            time.sleep(3)
            while GPIO.input(ir1) == 1:
                continue
            #servo_close()
            myservo.write(0)
            print("Gate Closed")
            time.sleep(3)
            count=count-1
            
            dist1=car_detection(slot1_trig,slot1_echo)
            dist2=car_detection(slot2_trig,slot2_echo)
            if (dist1<50):
                print("Car parked in slot-1")
                a=1
            else:
                print("No car parked in slot-1")
                a=0
            if (dist2<50):
                print("Car parked in slot-2")
                b=1
            else:
                print("No car parked in slot-2")
                b=0
            cloud(a,b,count)
            print("Number of cars parked =",count)
            time.sleep(15)
            
        
        if (GPIO.input(ir1)==0) and count>=2:
            print("Sorry!!! All slots Occupied..")
            
        time.sleep(2)




